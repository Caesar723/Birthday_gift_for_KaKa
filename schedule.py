from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont
import os
from datetime import datetime, timedelta
import threading

from functions import *
from rotate import *
from whether import Weather_Reporter




ORGPATH=os.path.dirname(os.path.abspath(__file__))
PATH=ORGPATH+"/{}"

FONT_SIZE=40
SPACE_FOR_LEFT=10


class Box:
    width_box=0
    height_box=0
    def __init__(self,x_index,y_index,text,draw,font) -> None:
        self.x=0 if x_index==0 else (x_index-1)*self.width_box+Box_time.width_box
        self.y=0 if y_index==0 else (y_index-1)*self.height_box+Box_day.height_box
        self.text=text
        self.draw=draw
        self.font=font
        self.highlight_area = (self.x, self.y, self.x+self.width_box, self.y+self.height_box)
        self.outline_color=(0,0,0)
    def draw_box(self):
        self.draw.rectangle(self.highlight_area, outline=self.outline_color, width=3)
        self.draw.text((self.x+SPACE_FOR_LEFT, self.y+self.height_box/2-FONT_SIZE/2), self.text, font=self.font, fill="black")


class Box_time(Box):
    width_box=280
    height_box=150
    def __init__(self, x_index, y_index, text, draw, font,whether_ana_time=True) -> None:
        super().__init__(x_index, y_index, text, draw, font)
        if whether_ana_time:
            self.time=0

    

class Box_content(Box):
    width_box=280
    height_box=150
    

class Box_day(Box):
    width_box=280
    height_box=150
   
class Table:
    img_width = 0#1600
    img_height = 0#1200 
    highlight_color = (131,194,233)

    def __init__(self,position,rotate_angle) -> None:
        self.position=position
        self.rotate_angle=rotate_angle
        
        #打开xlsx文件
        excel_file = PATH.format("files/schedule.xlsx")
        wb = load_workbook(excel_file)
        
        self.sheet = wb.active
        self.arr=list(self.sheet.iter_rows(values_only=True))
        
        #修改长度
        
        self.font = ImageFont.truetype(PATH.format("files/font.ttf"), FONT_SIZE, encoding="unic")
        
        self.filter_list()
        self.adjust_length()
        #初始化图片
        self.img = Image.new("RGBA", (self.img_width, self.img_height), "white")
        self.draw = ImageDraw.Draw(self.img)
        

       

        self.initinal_Box()
        self.high_light()
        self.draw_box()

        flipped_image = self.img.transpose(Image.FLIP_TOP_BOTTOM).transpose(Image.FLIP_LEFT_RIGHT)
        data = np.array(flipped_image)
        self.sch_textures=img2texture(data,(self.img_width, self.img_height))

        self.position_process()
        
    def physical_len(self,string):
        
        return int(self.font.getlength(string))
        
    def filter_list(self):
        
        self.titles_text=self.arr[0][1:]
        self.times_text=[row[0] for row in self.arr[1:]]
        self.contents_text=[[value if value!=None else "" for value in row[1:]] for row in self.arr[1:]]
        

    
    def adjust_length(self):
        max_content=max([max(content, key=self.physical_len) for content in self.contents_text], key=self.physical_len)
        Box_content.width_box=Box_day.width_box=self.physical_len(max_content)+40#len(max_content)*FONT_SIZE+40
        #最高400
        print(Box_content.width_box)

        max_time=max(self.times_text, key=self.physical_len)
        Box_time.width_box=self.physical_len(max_time)+40#len(max_time)*FONT_SIZE+40
        self.img_width=len(self.titles_text)*Box_content.width_box+Box_time.width_box
        self.img_height=len(self.times_text)*Box_content.height_box+Box_day.height_box

        
        
    def initinal_Box(self):
        self.titles=[Box_time(0,0,self.arr[0][0],self.draw,self.font,False)]+\
            [Box_day(i+1,0,self.titles_text[i],self.draw,self.font) for i in range(len(self.titles_text))]
        self.times=[Box_time(0,i+1,self.times_text[i],self.draw,self.font) for i in range(len(self.times_text))]
        self.contents=[
                [
                    Box_content(col+1,row+1,self.contents_text[row][col],self.draw,self.font) 
                    for col in range(len(self.contents_text[row]))
                ] 
                for row in range(len(self.contents_text))
            ]
        
    def draw_box(self):
        for title in self.titles:
            title.draw_box()
        for time in self.times:
            time.draw_box()
        for contents_row in self.contents:
            for content in contents_row:
                content.draw_box()


    def check_in_time_range(self,times:str):
        time1,time2=times.split("-")
        time1=datetime.strptime(time1, "%H:%M")
        time2=datetime.strptime(time2, "%H:%M")
        now=datetime.strptime(datetime.now().strftime('%H:%M'),"%H:%M")
        #print(now)
        if time1<now and now<time2:
            return True
        else:
            return False
        

    def high_light(self):
        #print(self.times_text)
        day_index={"Monday":0,"Tuesday":1,"Wednesday":2,"Thursday":3,"Friday":4,"Saturday":5,"Sunday":6}

        today = datetime.today()
        day_of_week = today.strftime("%A")
        i_day=day_index[day_of_week]

        for i in range(len(self.times_text)):
            if self.check_in_time_range(self.times_text[i]):
                self.draw.rectangle(self.contents[i][i_day].highlight_area, fill=self.highlight_color, width=3)

    def position_process(self):
        change_pos=to_np_array(self.position)
        four_corner=[(0.0, 0.0),(1.0, 0.0),(1.0, 1.0),(0.0, 1.0)]
        self.four_points=[[],[],[]]
        for corner in four_corner:
            x_unit=1 if corner[0]==1 else -1
            y_unit=1 if corner[1]==1 else -1
            self.four_points[0].append(self.position[0]+x_unit*self.img_width/100)
            self.four_points[1].append(self.position[1]+y_unit*self.img_height/100)
            self.four_points[2].append(0)
        self.four_points=np.array(self.four_points)
        self.four_points=Ry(self.rotate_angle)*self.four_points
        self.four_points=[self.four_points[:,i]+change_pos for i in range(4)]


    def create_weather_table(self):
        try:
            self.reporter=Weather_Reporter()
            thread=threading.Thread(target=self.reporter.get_answer_from_chatgpt)
            thread.start()

        except:
            print("Network Error")
        
    def check_weather_create(self):
        if self.reporter.answer!="":
            pass



            self.reporter.answer=""

            


    def display(self):
        #self.position=[0,0,0]
        glPushMatrix()
        glEnable(GL_TEXTURE_2D)
        #glEnable(GL_COLOR_MATERIAL)
        glDisable(GL_LIGHTING)
        glDisable(GL_LIGHT0)
        glEnable(GL_BLEND)

        glBlendFunc(GL_SRC_ALPHA, GL_ONE_MINUS_SRC_ALPHA)

        glColor4f(1,1,1,0.7)  # 设置颜色和透明度
        glBindTexture(GL_TEXTURE_2D, self.sch_textures)
        glBegin(GL_QUADS)

        four_corner=[(0.0, 0.0),(1.0, 0.0),(1.0, 1.0),(0.0, 1.0)]

        for point,corner in zip(self.four_points,four_corner):
            glTexCoord2f(*corner)
            
            glVertex3f(point[0][0], point[1][0],point[2][0])
        
        
        glEnd()
        glEnable(GL_LIGHTING)
        glEnable(GL_LIGHT0)
        glDisable(GL_TEXTURE_2D)
        glDisable(GL_BLEND)
        #glDisable(GL_COLOR_MATERIAL)
        glPopMatrix()

    







if __name__=="__main__":
    table=Table()
    
    #table.adjust_length()
    table.img.show()

