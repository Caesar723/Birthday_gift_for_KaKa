import requests
from datetime import datetime
import ipinfo
import os
from openpyxl import load_workbook
import json




ORGPATH=os.path.dirname(os.path.abspath(__file__))
PATH=ORGPATH+"/{}"



class Weather_Reporter:

    API_KEY_openWeather = "f077afc158a17acaef193a6f67b0f5d3"
    API_KEY_ipinfo="3dede6b592d7e4"
    

    
    
    def __init__(self) -> None:
        self.CITY = "Manchester"
        
        self.question=""
        self.answer=""
        

        

    def get_schedule_exal(self):
        excel_file=PATH.format("files/schedule.xlsx")
        wb = load_workbook(excel_file)
        sheet = wb.active
        arrs=list(sheet.iter_rows(values_only=True))

        day_index={"Monday":0,"Tuesday":1,"Wednesday":2,"Thursday":3,"Friday":4,"Saturday":5,"Sunday":6}
        today = datetime.today()
        day_of_week = today.strftime("%A")
        i_day=day_index[day_of_week]+1

        sche_today={}
        arrs.pop(0)
        for arr in arrs:
            sche_today[arr[0]]=arr[i_day]

        self.question+=str(sche_today)
        #print(str(sche_today))


    def forcast_today(self):
        URL = f"http://api.openweathermap.org/data/2.5/forecast?q={self.CITY}&appid={self.API_KEY_openWeather}"
        response = requests.get(URL)
        data = response.json()
        today=datetime.utcfromtimestamp(data["list"][0]["dt"]).day
        for forecast in data["list"]:
            timestamp = forecast["dt"]
            
            #print(datetime.utcfromtimestamp(timestamp).day)
            if datetime.utcfromtimestamp(timestamp).day!=30:
                break
            else:
                self.question+=f"{forecast}\n\n"
                #print(forecast)
                #print()

    def current_weather(self):
        URL = f"https://api.openweathermap.org/data/2.5/weather?q={self.CITY}&appid={self.API_KEY_openWeather}"
        response = requests.get(URL)
        data = response.json()
        #print()
        self.question+=f"{data}\n\n"
        #print(data)

    def get_current_city(self):
        handler = ipinfo.getHandler(self.API_KEY_ipinfo)
        
        details = handler.getDetails()
        return (details.city)
    
    

    
if __name__=="__main__":
    reporter=Weather_Reporter()
    reporter.get_answer_from_chatgpt()
